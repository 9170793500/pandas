
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# import pandas as pd
# import math as m
# import smtplib
# from email.message import EmailMessage
# from sqlalchemy import create_engine

# df = pd.read_csv('master.csv', header=None)
# columns = [3, 9]
# data = df.iloc[:, columns] #column select 

# ceiled_pulse = df.iloc[:, 13].astype(float).apply(lambda x: m.ceil(x) // 15)

# Equle_data = pd.concat([data, ceiled_pulse], axis=1)
# Equle_data.columns = ['User', 'Date', 'Pulse'] #make in hedder 

# df = pd.DataFrame(Equle_data)

# df['Date'] = pd.to_datetime(df['Date'])
# df['Date'] = df['Date'].dt.date

# Equle_data.insert(0, 'id', range(1, len(Equle_data) + 1))
# grouped = df.groupby(['User', 'Date'])['Pulse'].sum().reset_index() 
# grouped.columns = ['User','Date', 'Total_Pulse']


# print(grouped.head())

# grouped.to_excel("sum_by_date.xlsx", sheet_name="Sum_By_Date", index=False)

# # database connection
# database_uri = 'mysql+pymysql://root:@localhost/s'
# engine = create_engine(database_uri)
# query = "SELECT * FROM excel"
# try:
#   grouped.to_sql(name='excel', con=engine, if_exists='replace', index=False)
#   print("succes")  
# except:
#     print("not") 


# # Reading data from the Excel file
# excel_file = "sum_by_date.xlsx"
# sdm = pd.read_excel(excel_file)
# df = pd.DataFrame(sdm)

# # Function to color with HTML table 
# def table(df):
#     colors = {
#         'Date': '#FFFFE0',  # light yellow
#         'Total_Pulse': '#90EE90'  # light green
       
#     }
#     html_table = '<table style="border: 1px solid white; border-collapse: collapse;"><thead><tr>'
    
#     # Add table  in header with  color
#     for col in df.columns:
#         html_table += f'<th style="padding: 5px; background-color: {colors.get(col, "white")}">{col}</th>'
    
#     html_table += '</tr></thead><tbody>'
    
#     # Add table rows with column
#     for index, row in df.iterrows():
#         html_table += '<tr>'
#         for col in df.columns:
#             html_table += f'<td style="padding: 5px; background-color: {colors.get(col, "white")}">{row[col]}</td>'
#         html_table += '</tr>'
    
#     html_table += '</tbody></table>'
#     return html_table


# data_html = table(df)

# subject = "This is a check mail"
# body_template = f"""
# <html>
# <head>
# <style> 
#   table, th, td {{ border: 1px solid black; border-collapse: collapse; }}
#   th, td {{ padding: 5px; }}
# </style>
# </head>

# <body>
# {data_html}
# <br>
# Regards,
# Sundram Maurya
# </body>
# </html>
# """
# # Email in with HTML body
# msg = EmailMessage()
# msg['From'] = 'interns@fonada.com'
# # msg['To'] = 'mauryasundram1996@gmail.com'
# msg['To'] = 'ravindra@fonada.com'
# msg['Subject'] = subject
# msg.add_alternative(body_template, subtype='html')

# server = smtplib.SMTP('smtp.gmail.com', 587)
# # Login  account email and password
# server.starttls()
# server.login('sundrammaurya1996@gmail.com', 'okua lodv qjve oszg')  # gmail with your password

# with open("sum_by_date.xlsx", 'rb') as file:
#     file_data = file.read()
#     msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename='sum_by_date.xlsx')

# server.send_message(msg)
# server.quit()




from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import math as m
import smtplib
from email.message import EmailMessage
from sqlalchemy import create_engine

# Load CSV data
df = pd.read_csv('master.csv', header=None)

# Select specific columns
columns = [3, 9]
data = df.iloc[:, columns]

# Calculate ceiled pulse
ceiled_pulse = df.iloc[:, 13].astype(float).apply(lambda x: m.ceil(x) // 15)

# data with ceiled_pulse
Equle_data = pd.concat([data, ceiled_pulse], axis=1)
Equle_data.columns = ['User', 'DateTime', 'Pulse']  # Rename columns

# Create DataFrame
df = pd.DataFrame(Equle_data)

# Convert DateTime to datetime object and extract time
df['DateTime'] = pd.to_datetime(df['DateTime'])
df['Time'] = df['DateTime'].dt.strftime('%Y-%m-%d %H:%M:%S')  # Include date in 'Time' column

# Insert 'id' column
Equle_data.insert(0, 'id', range(1, len(Equle_data) + 1))

# Group by 'User' and 'Time' and calculate total pulse
grouped = df.groupby(['User', 'Time'])['Pulse'].sum().reset_index()
grouped.columns = ['User', 'Time', 'Total_Pulse']

print(grouped.head())

# Save grouped data to Excel file
grouped.to_excel("sum_by_time.xlsx", sheet_name="Sum_By_Time", index=False)

# Database connection
database_uri = 'mysql+pymysql://root:@localhost/s'
engine = create_engine(database_uri)
query = "SELECT * FROM excel"

#  replace the data  'excel' table
try:
    grouped.to_sql(name='excel', con=engine, if_exists='replace', index=False)
    print("Success")
except:
    print("Error")

# Read data from the Excel file
excel_file = "sum_by_time.xlsx"
stm = pd.read_excel(excel_file)
df = pd.DataFrame(stm)

# Function to create HTML table
def table(df):
    colors = {
        'Time': '#FFFFE0',  # light yellow
        'Total_Pulse': '#90EE90'  # light green
    }

    html_table = '<table style="border: 1px solid white; border-collapse: collapse;"><thead><tr>'

    # Add table header with color
    for col in df.columns:
        html_table += '<th style="padding: 5px; background-color: ' + colors.get(col, "white") + '">' + col + '</th>'

    html_table += '</tr></thead><tbody>'

    # Add table rows with column values
    for index, row in df.iterrows():
        html_table += '<tr>'
        for col in df.columns:
            html_table += '<td style="padding: 5px; background-color: ' + colors.get(col, "white") + '">' + str(row[col]) + '</td>'
        html_table += '</tr>'

    html_table += '</tbody></table>'
    return html_table

# Generate HTML table
data_html = table(df)

# Email configuration
subject = "Pulse Data Report"
body_template = """
<html>
<head>
<style> 
  table, th, td { border: 1px solid black; border-collapse: collapse; }
  th, td { padding: 5px; }
</style>
</head>

<body>
""" + data_html + """
<br>
Date: """ + str(df['Time'].iloc[0]) + """  # Use 'Time' instead of 'DateTime' in the email body
<br>
Regards,
Sundram Maurya
</body>
</html>
"""

# Email setup with HTML body
msg = EmailMessage()
msg['From'] = 'sundrammaurya1996@gmail.com'
msg['To'] = 'vipin140.chauhan@gmail.com'
msg['Subject'] = subject
msg.add_alternative(body_template, subtype='html')

# Attach the Excel file to the email
with open("sum_by_time.xlsx", 'rb') as file:
    file_data = file.read()
    msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename='sum_by_time.xlsx')

# SMTP server configuration and email sending
server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login('sundrammaurya1996@gmail.com','okua lodv qjve oszg')  # Gmail login credentials

# Send the email
server.send_message(msg)
server.quit()
